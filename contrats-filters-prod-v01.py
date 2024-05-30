import yaml
from openpyxl import load_workbook




class MyDumper(yaml.Dumper):

    def increase_indent(self, flow=False, indentless=False):
        return super(MyDumper, self).increase_indent(flow, False)
# Load Excel file
workbook = load_workbook(filename='ListeContratEtFiltresV15.xlsx')
sheet = workbook['ContratToFilters']
# Initialize data structure to hold the YAML content
data = {"apic": {"tenants": []}}
current_tenant = None
# Iterate through each row in the Excel sheet
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Extract values from the row
    tenant_name, contract_name, subject_name, *filters_names = row
    filters_names = [filter_name for filter_name in filters_names if filter_name != None]
    # Construct the YAML structure

    if current_tenant != tenant_name:
        current_tenant = tenant_name
        data["apic"]["tenants"].append({"name": tenant_name, "contracts": []})
    data["apic"]["tenants"][-1]["contracts"].append({"name": contract_name,"subjects": [{"name": subject_name, "filters": []}]})
    # Append each filter to the filters list under the subject
    for filter_name in filters_names:
        data["apic"]["tenants"][-1]["contracts"][-1]["subjects"][-1]["filters"].append({"filter": filter_name})

# Write the YAML content to a file
    # Append the tenant data to the main data structure


# Write the YAML content to a file
with open("contract-filters.yaml", "w") as yaml_file:
    yaml.dump(data, yaml_file, Dumper=MyDumper, default_flow_style=False,sort_keys=False, allow_unicode=True)

print("YAML file generated successfully!")
