import yaml
from openpyxl import load_workbook



class MyDumper(yaml.Dumper):

    def increase_indent(self, flow=False, indentless=False):
        return super(MyDumper, self).increase_indent(flow, False)
# Load the Excel file
workbook = load_workbook('FiltresTests.xlsx')
sheet = workbook['Unique_Filters']

data = {"apic": {"tenants": []}}
current_tenant = None
current_filter = None
current_epg = None
dst_to_port = None


for row in sheet.iter_rows(min_row=2, values_only=True):
    tenant_name, filter_name, entry_name, ethertype, protocole, dst_from_port, dst_to_port, stateful = row
   
    if current_tenant != tenant_name:
        current_tenant = tenant_name
        data["apic"]["tenants"].append({"name": tenant_name, "filters": []})

    if current_filter != filter_name:
        current_filter = filter_name
        data["apic"]["tenants"][-1]["filters"].append({"name": filter_name,"entries": []})
    
    if dst_to_port != None :  
        data["apic"]["tenants"][-1]["filters"][-1]["entries"].append({"name": entry_name, "ethertype": ethertype, "protocol":protocole, "destination_from_port":dst_from_port, "destination_to_port":dst_to_port, "stateful": stateful})
    else:
         data["apic"]["tenants"][-1]["filters"][-1]["entries"].append({"name": entry_name, "ethertype": ethertype, "protocol":protocole, "destination_from_port":dst_from_port, "stateful": stateful})

# Write the data to a YAML file
with open("filters.yaml", "w") as yaml_file:
    yaml.dump(data, yaml_file, Dumper=MyDumper, default_flow_style=False,sort_keys=False)