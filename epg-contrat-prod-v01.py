import yaml
from openpyxl import load_workbook


class MyDumper(yaml.Dumper):

    def increase_indent(self, flow=False, indentless=False):
        return super(MyDumper, self).increase_indent(flow, False)

workbook = load_workbook(filename='EPGsContrats.xlsx')
sheet = workbook['Contrat_EPGs']

# Initialize data structure to hold the YAML content
data = {"apic": {"tenants": []}}
current_tenant = None
for row in sheet.iter_rows(min_row=2, values_only=True):
    # Extract values from the row
    tenant_name, contract_name, app_profile_name, contract_name, epg_source, epg_destination = row
    print(epg_destination)
    if current_tenant != tenant_name:
        current_tenant = tenant_name
        data["apic"]["tenants"].append({"name": tenant_name, "application_profiles": [{"name":app_profile_name,"endpoint_groups": []} ]})
    data["apic"]["tenants"][-1]["application_profiles"][-1]["endpoint_groups"].append({
        'name': epg_source,
        'contracts': {
            'consumers': [contract_name]
        }
    })
    if epg_source != epg_destination:
        data["apic"]["tenants"][-1]["application_profiles"][-1]["endpoint_groups"].append({
        'name': epg_destination,
        'contracts': {
            'providers': [contract_name]
        }
    })
# Initialize data structure to hold the YAML content

# Initialize the YAML data structure

# Save the YAML data to a file

# Write the YAML content to a file
with open('epg-contract.yaml', 'w') as yaml_file:
    yaml.dump(data, yaml_file, Dumper=MyDumper, default_flow_style=False,sort_keys=False, allow_unicode=True)

print("YAML file generated successfully!")
