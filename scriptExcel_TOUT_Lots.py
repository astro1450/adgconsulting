from openpyxl import load_workbook
import openpyxl
def unmerge_and_fill(workbook_path,unmerged_file):
    # Load the Excel workbook
    wb = load_workbook(workbook_path)
    
    # Select the active sheet
    ws = wb.active
    
    # Create a copy of the set of merged cell ranges
    merged_cell_ranges = set(ws.merged_cells.ranges.copy())
    
    # Loop through all merged cells in the worksheet
    for merged_cell_range in merged_cell_ranges:
        # Unmerge the cells
        ws.unmerge_cells(str(merged_cell_range))
  
        # Extract the value of the merged cell
        merged_value = ws.cell(row=merged_cell_range.min_row, column=merged_cell_range.min_col).value
        
        # Fill the content of the merged cell into all the unmerged cells
        for row in range(merged_cell_range.min_row, merged_cell_range.max_row + 1):
            for col in range(merged_cell_range.min_col, merged_cell_range.max_col + 1):
                if row != merged_cell_range.min_row or col != merged_cell_range.min_col:
                    ws.cell(row=row, column=col).value = merged_value
    
    # Save the changes
    wb.save(unmerged_file)

# Example usage:
unmerged_file = "unmerged_file1.xlsx" 
workbook_path = "MatriceFlux_LOT-ADEHT_PROD.xlsx"
unmerge_and_fill(workbook_path,unmerged_file)
# Example usage:
def delete_first_rows_excel(input_file_path, output_file_path, num_rows):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file_path)
    sheet = wb.active
    sheet.views.sheetView[0] = openpyxl.worksheet.views.SheetView()
    sheet.auto_filter.ref = None


    # Delete the specified number of rows
    for _ in range(num_rows):
        sheet.delete_rows(1)
    
    # Save the modified workbook to a new file
    wb.save(output_file_path)


rows_deleted_file = "rows_deleted_file2.xlsx"


num_rows_to_delete = 10
delete_first_rows_excel(unmerged_file, rows_deleted_file, num_rows_to_delete)


from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

def delete_columns(input_file , columns, output_file):
    # Load the Excel workbook
    wb = load_workbook(input_file)
    
    # Iterate over each worksheet in the workbook
    for ws in wb.worksheets:
        # Convert column letters to numerical indices
        column_indices = [column_index_from_string(col) for col in columns]
        
        # Sort column indices in descending order to avoid index shifting
        column_indices.sort(reverse=True)
        
        # Delete columns in descending order to avoid shifting indices
        for col_idx in column_indices:
            ws.delete_cols(col_idx)
    
    # Save the changes
    wb.save(output_file)

# Example usage:

columns_deleted_file = "columns_deleted_file3.xlsx"
columns_to_delete = ["C", "J", "K"]
delete_columns(rows_deleted_file, columns_to_delete,columns_deleted_file)



import openpyxl

def find_and_replace1(file_path, sheet_name, column_let, old_value1, old_value2, new_value):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Iterate over all the rows and columns in the sheet
    for cell in sheet[column_let]:
        if cell.value and not isinstance(cell.value, int):
            if cell.value and old_value1 in cell.value :
                cell.value = cell.value.replace(old_value1, new_value)
            elif cell.value and old_value2 in cell.value :
                cell.value = cell.value.replace(old_value2, new_value)
            elif  cell.value and "Type" in cell.value :
                cell.value = "ICMP"

    # Save the workbook
    workbook.save(file_path)

# Usage
file_path = columns_deleted_file # Replace with the path to your Excel file
sheet_name = 'Matrice de flux'       # Replace with the name of your sheet
old_value1 = 'XDA_USERS_VRF'     # Replace with the value you want to find
old_value2 = 'XDA_USERS'
new_value = 'XDA_DATA_VRF'     # Replace with the value you want to replace with
column_let= 'E'
find_and_replace1(file_path, sheet_name, column_let, old_value1, old_value2, new_value)
def find_and_replace2(file_path, sheet_name, column_let, old_value1, new_value):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Iterate over all the rows and columns in the sheet
    for cell in sheet[column_let]:
        if cell.value and not isinstance(cell.value, int):
            if cell.value and old_value1 in cell.value :
                cell.value = new_value

    # Save the workbook
    workbook.save(file_path)

# Usage
file_path = columns_deleted_file # Replace with the path to your Excel file
sheet_name = 'Matrice de flux'       # Replace with the name of your sheet


find_and_replace2(file_path, sheet_name, 'F', '_VLAN', 'ExternalEPG_Borders')











from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator

def add_headers_and_vrf_column(file_path, headers,added_headers_file):
    # Load the Excel workbook
    wb = load_workbook(file_path)
    
    # Iterate over each worksheet in the workbook
    for ws in wb.worksheets:
        # Determine the column letter for the new column
   
        max_column = ws.max_column
        new_column_letter = get_column_letter(max_column + 1)
        
        # Shift existing data down to make room for the new header
        ws.insert_rows(1)
        
        # Set the header for the new column
        ws[f"{new_column_letter}1"] = "VRF"
        
        # Set headers for all existing columns
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            ws[f"{col_letter}1"] = header
        
        # Iterate over rows to apply the formula
   # Iterate over rows to apply the formula
        for row in range(2, ws.max_row + 1):
            cell_address = f"{new_column_letter}{row}"
            # Set the formula for the cell
            vrf_formula = (
                '=IF(OR(C{row}="XDA_DATA_VRF", C{row}="XDA_ADM_VRF"), '
                'C{row}, '
                'IF(OR(E{row}="XDA_DATA_VRF", E{row}="XDA_ADM_VRF"), '
                'E{row}, '
                '"NoContracts"))'.format(row=row)
            )
            # Translate the formula to Python syntax
            ws[f"{new_column_letter}{row}"] = vrf_formula
    # Save the changes
    wb.save(added_headers_file)

# Example usage:



# Example usage:
added_headers_file = "added_headers_file4.xlsx"
file_path = columns_deleted_file
headers = ["","Id", "Src VRF", "EPG source", "Dst VRF","EPG destination", "all Ports", "Protocol"]  # Replace with your desired headers
add_headers_and_vrf_column(file_path, headers,added_headers_file)


#################################################""
import xlwings as xw
import openpyxl

def recalculate_and_save(file_path):
    """
    Recalculates all formulas in an Excel file and saves it.
    :param file_path: Path to the input Excel file.
    """
    app = xw.App(visible=False)
    wb = xw.Book(file_path)
    wb.app.calculate()
    wb.save()
    wb.close()
    app.quit()

import openpyxl

def transform_formulas_to_values(file_path, sheet_name, column_letter, output_file_path=None):
    """
    Transforms formulas to their resulting values in a specified column of an Excel sheet.

    :param file_path: Path to the input Excel file.
    :param sheet_name: Name of the sheet to process.
    :param column_letter: Letter of the column to transform formulas to values.
    :param output_file_path: Path to save the modified Excel file (default is to overwrite the input file).
    """
    recalculate_and_save(file_path)
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]

    # Load the Excel file with formulas intact
    workbook_with_formulas = openpyxl.load_workbook(file_path)
    sheet_with_formulas = workbook_with_formulas[sheet_name]

    # Iterate through the specified column and replace formulas with their values
    for row in range(1, sheet.max_row + 1):
        cell = f'{column_letter}{row}'
        formula_cell = sheet_with_formulas[cell]
        value_cell = sheet[cell]
        
        # Check if the cell contains a formula
        if formula_cell.data_type == 'f':  # 'f' means the cell contains a formula
            value_cell.value = value_cell.value

    # Save the modified Excel file
    if output_file_path is None:
        output_file_path = file_path
    workbook.save(output_file_path)




# Example usage
transform_formulas_to_values(
    file_path=added_headers_file,
    sheet_name='Matrice de flux',
    column_letter='I',
    output_file_path='toValues_excel_file.xlsx'  # Optional: specify a different output file path
)

# Example usage



######################################################################################

import openpyxl

def delete_rows_by_column_value(file_path, sheet_name, column_letter, value_to_delete, output_file_path):
    """
    Deletes rows from an Excel sheet based on a specified column value.

    :param file_path: Path to the input Excel file.
    :param sheet_name: Name of the sheet to process.
    :param column_letter: Letter of the column to check for the value.
    :param value_to_delete: Value that, if found in the column, will result in the row being deleted.
    :param output_file_path: Path to save the modified Excel file.
    """
    # Load the Excel file
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Iterate through the rows and collect the indices of the rows to delete
    rows_to_delete = []
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=openpyxl.utils.cell.column_index_from_string(column_letter), max_col=openpyxl.utils.cell.column_index_from_string(column_letter)):
        for cell in row:
            if cell.value == value_to_delete:
                rows_to_delete.append(cell.row)

    # Delete the rows from the bottom to the top
    for row_index in sorted(rows_to_delete, reverse=True):
        sheet.delete_rows(row_index)

    # Save the modified Excel file
    workbook.save(output_file_path)

# Example usage
delete_rows_by_column_value(
    file_path="toValues_excel_file.xlsx",
    sheet_name='Matrice de flux',
    column_letter='I',
    value_to_delete='NoContracts',
    output_file_path='NoContracts_excel_file.xlsx'
)


def delete_columns(input_file , columns, output_file):
    # Load the Excel workbook
    wb = load_workbook(input_file)
    
    # Iterate over each worksheet in the workbook
    for ws in wb.worksheets:
        # Convert column letters to numerical indices
        column_indices = [column_index_from_string(col) for col in columns]
        
        # Sort column indices in descending order to avoid index shifting
        column_indices.sort(reverse=True)
        
        # Delete columns in descending order to avoid shifting indices
        for col_idx in column_indices:
            ws.delete_cols(col_idx)
    
    # Save the changes
    wb.save(output_file)

# Example usage:

columns_deleted_file = "vrf_excel_file.xlsx"
columns_to_delete = ["C", "E"]
delete_columns("NoContracts_excel_file.xlsx", columns_to_delete,columns_deleted_file)







import openpyxl

def move_column(file_path, sheet_name, column_letter, new_position):
    """
    Moves a column in an Excel sheet to a new position.

    :param file_path: Path to the input Excel file.
    :param sheet_name: Name of the sheet to process.
    :param column_letter: Letter of the column to move.
    :param new_position: New position for the column (1-based index).
    """
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Convert column letter to column index
    column_index = openpyxl.utils.column_index_from_string(column_letter)

    # Read the entire column
    column_data = [cell.value for cell in sheet[column_letter]]

    # Insert a new empty column at the desired new position
    sheet.insert_cols(new_position)

    # Calculate the target column letter based on the new position
    target_column_letter = openpyxl.utils.get_column_letter(new_position)

    # Write the data to the new column position
    for row_num, value in enumerate(column_data, start=1):
        sheet[f'{target_column_letter}{row_num}'].value = value

    # Delete the original column
    if new_position <= column_index:
        # If moving left, original column index has increased by 1
        sheet.delete_cols(column_index + 1)
    else:
        sheet.delete_cols(column_index)

    # Save the modified Excel file
    workbook.save(file_path)

# Example usage
move_column(
    file_path='vrf_excel_file.xlsx',
    sheet_name='Matrice de flux',
    column_letter='G',  # Column to move
    new_position=3  # New position (1-based index)
)


import openpyxl

def find_and_replace(file_path, sheet_name, column_let, old_value1, old_value2, new_value):
    # Load the workbook and select the specified sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    
    # Iterate over all the rows and columns in the sheet
    for cell in sheet[column_let]:
        if cell.value and not isinstance(cell.value, int):
            if cell.value and old_value1 in cell.value :
                cell.value = cell.value.replace(old_value1, new_value)
            elif cell.value and old_value2 in cell.value :
                cell.value = cell.value.replace(old_value2, new_value)
            elif  cell.value and "Type" in cell.value :
                cell.value = "ICMP"

    # Save the workbook
    workbook.save(file_path)

# Usage
file_path = 'vrf_excel_file.xlsx'  # Replace with the path to your Excel file
sheet_name = 'Matrice de flux'       # Replace with the name of your sheet
old_value1 = ' ... '     # Replace with the value you want to find
old_value2 = ' .. '
new_value = '-'     # Replace with the value you want to replace with
column_let= 'F'
find_and_replace(file_path, sheet_name, column_let, old_value1, old_value2, new_value)


import openpyxl

def concatenate_columns_with_underscore(file_path, sheet_name, col1_letter, col2_letter, result_col_letter):
    """
    Concatenates values from two columns with an underscore between them and puts the result in a third column in an Excel sheet.

    :param file_path: Path to the input Excel file.
    :param sheet_name: Name of the sheet to process.
    :param col1_letter: Letter of the first column to concatenate.
    :param col2_letter: Letter of the second column to concatenate.
    :param result_col_letter: Letter of the column to store the concatenated results.
    """
    # Load the workbook and select the sheet
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Iterate through each row in the specified columns
    for row in range(2, sheet.max_row + 1):
        cell1 = sheet[f'{col1_letter}{row}']
        cell2 = sheet[f'{col2_letter}{row}']
        result_cell = sheet[f'{result_col_letter}{row}']

        # Concatenate the values with an underscore and write the result
        concatenated_value = f"{cell1.value or ''}_{cell2.value or ''}"
        result_cell.value = concatenated_value
    sheet[f'{result_col_letter}{1}'].value = "Ports"
    # Save the modified Excel file
    workbook.save(file_path)
# Example usage
concatenate_columns_with_underscore(
    file_path='vrf_excel_file.xlsx',
    sheet_name='Matrice de flux',
    col1_letter='F',  # First column to concatenate
    col2_letter='G',  # Second column to concatenate
    result_col_letter='H'  # Column to store the concatenated results
)

move_column(
    file_path='vrf_excel_file.xlsx',
    sheet_name='Matrice de flux',
    column_letter='B',  # Column to move
    new_position=9  # New position (1-based index)
)

